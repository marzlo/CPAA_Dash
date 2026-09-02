import openpyxl, re, json, csv, os, datetime as _dt

# In the GitHub Actions pipeline, fetch_jira.py writes this CSV (same column shape as a
# manual Jira CSV export) right before this script runs. Override with JIRA_CSV_PATH if needed.
SRC = os.environ.get("JIRA_CSV_PATH", "jira_export.csv")

# Support both the .xlsx export (openpyxl) and the raw Jira .csv export (csv module),
# normalising both into the same header/data shape so the rest of the script is unchanged.
if SRC.lower().endswith(".csv"):
    with open(SRC, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    header = rows[0]
    data = [tuple((cell if cell != "" else None) for cell in row) for row in rows[1:]]
else:
    wb = openpyxl.load_workbook(SRC, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if s.startswith("CPAA_general_ticket")), wb.sheetnames[0])
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data = rows[1:]

# Resolve column positions by header name (Jira's export column order/count shifts
# between exports as fields get added, so never hardcode positional indices).
IDX_TYPE = header.index("Issue Type")
IDX_KEY = header.index("Issue key")
IDX_SUMMARY = header.index("Summary")
IDX_STATUS = header.index("Status")
IDX_ASSIGNEE = header.index("Assignee") if "Assignee" in header else None
IDX_SEVERITY = header.index("Custom field (Severity)") if "Custom field (Severity)" in header else None
IDX_PRIORITY = header.index("Priority") if "Priority" in header else None
IDX_CREATED = header.index("Created") if "Created" in header else None
LABEL_COLS = [i for i, h in enumerate(header) if h == "Labels" or (isinstance(h, str) and h.startswith("Labels_"))]

# Normalise the "Created" cell to an ISO date string regardless of source: the CSV path
# (fetch_jira.py, live pipeline) gives an ISO datetime string straight from the Jira API;
# the manual .xlsx export path gives an openpyxl-parsed datetime/date object.
def created_iso(v):
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.isoformat()
    return v or None

# Assignee -> team/org mapping, sourced from the cpaa-dashboard skill's TEAM_MAP
# (build_all_swe.py). Anyone not in here falls under "Unknown" and should be
# flagged to the user rather than silently guessed.
TEAM_MAP = {
    '孙文奇': 'TS_FW', '孙琦': 'TS_CPAA', 'zhaoyuchen': 'TS_FW', 'Unassigned': 'Unassigned',
    'JesseCHHuang(黃昭華)': 'MDT_System', 'KVVD Subrahmanyam': 'MDI_System',
    'FrankYHYang(楊宇翔)': 'MDT_System', 'LeyoYYLin(林沅佑)': 'MDT_PM',
    'SHANE YEH': 'MDT_App', 'Sudha M': 'MDI_System', 'Santoshkumar Hiremath': 'MDI_System',
    'VincentWang(王志玄)': 'MDT_System', 'CCCJHuang(黃清俊)': 'MDT_System',
    'BrianYang': 'MDT_App', 'DeanDYJiang(江定遠)': 'MDT_System',
    'StanleyKCWu(吳冠麒)': 'MDT_App', 'yitong.xu_b': 'TS_CPAA',
    'Sithala Kalyan Kumar': 'MDI_System', '姜馨雨': 'TS_CPAA',
    'RayYCHsieh(謝逸群)': 'MDT_App', 'Justin Chen': 'MDT_App',
    'Weili Chou(周洧立)': 'MDT_System',
    'jinzhe.wang': 'TS_FW', 'ShawnHHLee(李湘豪)': 'MDT_System', 'JohnnyHJ Lin': 'MDT_App',
    'yu_yong': 'TS_CPAA', '于泳': 'TS_CPAA',
    # Added per user confirmation (2026-08-21)
    'HuanYangLin(林奐揚)': 'MDT_System', 'Samarsinh': 'MDI_System',
    'StevenSCChang(張舜喬)': 'MDT_App', '何明轩': 'TS_CPAA', '刘显鹤': 'TS_FW',
    '唐竞远': 'TS_CPAA', '孟庆奎': 'TS_CPAA', '张立双(ZhangLishuang)': 'TS_CPAA',
    '杜鸿运': 'TS_CPAA', '王锦喆': 'TS_FW', '辛颖': 'TS_CPAA',
    'Subramanya N S': 'MDI_System', 'StevenSFYang': 'MDT_PM',
    'NikKMLiu': 'MDT_App', 'EricYou(游超雲)': 'MDT_System',
    'StevenJSHsu(許智舜)': 'MDT_System', '商玉函': 'TS_CPAA',
    'TS-赵洪': 'TS_FW',
    'JerryCCChen(陳璟錞)': 'MDT_PM',
    '杨百全': 'TS_CPAA',
    # Added per user confirmation (2026-09-01)
    'Shaik Mohammed Arif': 'MDI_System',
}

DONE_STATUSES = {"Done"}  # statusCategory 'done'
# Jira statusCategory mapping (approx from status names seen)
NEW_STATUSES = {"To Do", "Blocked", "Reopen", "Need info"}
INDET_STATUSES = {"In Progress", "Ready for review", "Ready for integration"}
# "Non-issue" -> treat as done-ish (closed, won't fix) -> count as done for completion purposes? We'll bucket separately.

# Per user request: for Bug tickets only, "Ready for test", "Eng build" and "Monitoring" also count as completed.
BUG_EXTRA_DONE_STATUSES = {"Ready for test", "Eng build", "Monitoring"}

def status_category(status, extra_done=None):
    if status in DONE_STATUSES:
        return "done"
    if status == "Non-issue":
        return "done"  # closed/resolved, excluded from "not completed"
    if extra_done and status in extra_done:
        return "done"
    if status in NEW_STATUSES:
        return "new"
    if status in INDET_STATUSES:
        return "indeterminate"
    return "new"

def labels_of(r):
    return sorted(set(r[i] for i in LABEL_COLS if r[i]))

def classify_feature(summary):
    s = (summary or "").strip()
    if re.match(r'^ipod', s, re.I):
        return "iPod"
    if re.match(r'^cp\b', s, re.I) or s.upper().startswith("CP["):
        return "CarPlay"
    if re.match(r'^aa\b', s, re.I) or s.upper().startswith("AA["):
        return "Android Auto"
    return "Other"

def classify_feature_bug(summary):
    s = summary or ""
    if re.search(r'ipod', s, re.I):
        return "iPod"
    if re.search(r'android[\s\-]?auto', s, re.I) or re.search(r'\bAA\b', s):
        return "Android Auto"
    if re.search(r'carplay', s, re.I) or re.search(r'\bCP\b', s):
        return "CarPlay"
    return "Other"

def label_bucket(has_r2, has_r3, has_cpaa0830):
    if has_r2:
        return "ASW-R2"
    if has_cpaa0830:
        return "CPAA0830"
    if has_r3:
        return "ASW-R3 (不含CPAA 0830)"
    return "三者皆無"

# --- CarPlay / Android Auto / iPod sub-feature taxonomy (per cpaa-feature-taxonomy skill) ---
# Best-effort keyword classification against the fixed sub-feature list. Tickets that don't
# clearly match any sub-feature are left as None ("未分類") rather than force-guessed.
_GROUP_PREFIX = {"CarPlay": "CarPlay", "Android Auto": "Android Auto", "iPod": "IPod"}

def classify_subfeature(feature, summary):
    g = _GROUP_PREFIX.get(feature)
    if g is None:
        return None
    sl = (summary or "").lower()

    def has(*words):
        return any(w in sl for w in words)

    # --- Manually confirmed classifications (user-specified, 2026-08-21) ---
    if has("keymasterwrapper"):
        return f"{g}_Pre-certification"
    if has("dr algorithm", "gnss requirement"):
        return f"{g}_Navigation"
    if has("pascd location"):
        return f"{g}_Interaction with Local Reverse Gear"
    if has("power interaction", "power shutdown", "power str"):
        return f"{g}_Power"
    if has("hardware and software version api"):
        return f"{g}_Other"
    if has("reconnection logic architecture"):
        return f"{g}_Wireless Disconnect/Reconnect"
    if has("pre-certification"):
        return f"{g}_Pre-certification"

    if has("reverse gear", "avm reverse"):
        return f"{g}_Interaction with Local Reverse Gear"
    if has("instrument cluster"):
        return f"{g}_Instrument Cluster Display"
    if has("local media display", "launcher display", "widget"):
        return f"{g}_Local Media Display"
    if has("device switching", "device management", "favorite device"):
        return f"{g}_Device Switching"
    if has("siri", "esiri"):
        return f"{g}_Enhance Siri" if "enhance" in sl else f"{g}_Classic Siri"
    if has("hardkey", "hard key", "rotary knob", "steering wheel", "screen-off button"):
        return f"{g}_Hardkey"
    if feature == "Android Auto" and has(" vr ", "vr)", "(vr", "voice recognition"):
        return f"{g}_VR"
    if has("navigation metadata", "navigation"):
        return f"{g}_Navigation"
    if has("phone", "call", "facetime") and "audio" not in sl:
        if has("facetime"):
            return f"{g}_FaceTime"
        return f"{g}_Phone"
    if has("music") and "audio" not in sl:
        return f"{g}_Music"
    if has("wireless", "wifi", "wi-fi", "oob", "bluetooth", "hci", "pairing"):
        if has("disconnect", "reconnect", "reconnection"):
            return f"{g}_Wireless Disconnect/Reconnect"
        return f"{g}_Wireless Connection"
    if has("wired"):
        if has("disconnect", "reconnect"):
            return f"{g}_Wired Disconnect/Reconnect"
        return f"{g}_Wired Connection"
    if has("video", "display", "screensaver"):
        return f"{g}_Display"
    if has("audio", "ecnr", "aaudio", "】-【fm】"):
        return f"{g}_Interaction with Local Audio Source"
    if feature == "iPod":
        if has("media library", "jump list", "media display", "media function"):
            return f"{g}_Media Display"
        if has("playback"):
            return f"{g}_Playback Control"
    return None

records = []
for r in data:
    summary = r[IDX_SUMMARY] or ""
    swe = None
    if "SWE2" in summary:
        swe = "SWE2"
    elif "SWE3" in summary:
        swe = "SWE3"
    elif "SWE5" in summary:
        swe = "SWE5"
    else:
        continue
    labels = labels_of(r)
    status = r[IDX_STATUS]
    cat = status_category(status)
    assignee = (r[IDX_ASSIGNEE] if IDX_ASSIGNEE is not None else None) or "Unassigned"
    rec = {
        "key": r[IDX_KEY],
        "summary": summary,
        "status": status,
        "statusCategory": cat,
        "labels": labels,
        "swe": swe,
        "feature": classify_feature(summary),
        "done": cat == "done",
        "hasR2": "ASW-R2" in labels,
        "hasR3": "ASW-R3" in labels,
        "hasCPAA0830": "CPAA_0830" in labels,
        "assignee": assignee,
        "team": TEAM_MAP.get(assignee, "Unknown"),
        "created": created_iso(r[IDX_CREATED]) if IDX_CREATED is not None else None,
    }
    rec["labelBucket"] = label_bucket(rec["hasR2"], rec["hasR3"], rec["hasCPAA0830"])
    rec["subFeature"] = classify_subfeature(rec["feature"], summary) or "未分類"
    records.append(rec)

print("Total SWE2+SWE3+SWE5:", len(records))
swe2 = [r for r in records if r["swe"] == "SWE2"]
swe3 = [r for r in records if r["swe"] == "SWE3"]
swe5 = [r for r in records if r["swe"] == "SWE5"]
print("SWE2:", len(swe2), "done:", sum(r["done"] for r in swe2))
print("SWE3:", len(swe3), "done:", sum(r["done"] for r in swe3))
print("SWE5:", len(swe5), "done:", sum(r["done"] for r in swe5))
print("ASW-R2:", sum(r["hasR2"] for r in records))
print("ASW-R3:", sum(r["hasR3"] for r in records))
not_done_no_labels = [r for r in records if not r["done"] and not r["hasR2"] and not r["hasR3"]]
print("Not done & missing both labels:", len(not_done_no_labels))

not_done = [r for r in records if not r["done"]]
bucket_counts = {}
for r in not_done:
    bucket_counts[r["labelBucket"]] = bucket_counts.get(r["labelBucket"], 0) + 1
print("Not-done label buckets:", bucket_counts)

subfeature_counts = {}
for r in not_done:
    subfeature_counts[r["subFeature"]] = subfeature_counts.get(r["subFeature"], 0) + 1
print("Not-done sub-feature counts:", dict(sorted(subfeature_counts.items(), key=lambda x: -x[1])))
uncat = [r for r in not_done if r["subFeature"] == "未分類"]
print(f"Not-done 未分類 (no sub-feature match, {len(uncat)} tickets):")
for r in uncat:
    print("  ", r["key"], "|", r["swe"], "|", r["feature"], "|", r["summary"][:90])

feat_counts = {}
for r in records:
    feat_counts[r["feature"]] = feat_counts.get(r["feature"], 0) + 1
print("Feature counts:", feat_counts)

team_counts = {}
for r in not_done:
    team_counts[r["team"]] = team_counts.get(r["team"], 0) + 1
print("Not-done team counts:", dict(sorted(team_counts.items(), key=lambda x: -x[1])))
unknown_assignees = sorted(set(r["assignee"] for r in not_done if r["team"] == "Unknown"))
if unknown_assignees:
    print(f"Assignees with NO team mapping ({len(unknown_assignees)}), flagged as 'Unknown':")
    for a in unknown_assignees:
        print("  ", a)

# --- Bug tickets (Issue Type = Bug) ---
bug_records = []
for r in data:
    if r[IDX_TYPE] != "Bug":
        continue
    summary = r[IDX_SUMMARY] or ""
    labels = labels_of(r)
    status = r[IDX_STATUS]
    cat = status_category(status, extra_done=BUG_EXTRA_DONE_STATUSES)
    assignee = (r[IDX_ASSIGNEE] if IDX_ASSIGNEE is not None else None) or "Unassigned"
    severity = (r[IDX_SEVERITY] if IDX_SEVERITY is not None else None) or "未標示"
    priority = (r[IDX_PRIORITY] if IDX_PRIORITY is not None else None) or "未標示"
    rec = {
        "key": r[IDX_KEY],
        "summary": summary,
        "status": status,
        "statusCategory": cat,
        "labels": labels,
        "feature": classify_feature_bug(summary),
        "done": cat == "done",
        "hasR2": "ASW-R2" in labels,
        "hasR3": "ASW-R3" in labels,
        "hasCPAA0830": "CPAA_0830" in labels,
        "assignee": assignee,
        "team": TEAM_MAP.get(assignee, "Unknown"),
        "severity": severity,
        "priority": priority,
        "created": created_iso(r[IDX_CREATED]) if IDX_CREATED is not None else None,
    }
    rec["labelBucket"] = label_bucket(rec["hasR2"], rec["hasR3"], rec["hasCPAA0830"])
    rec["subFeature"] = classify_subfeature(rec["feature"], summary) or "未分類"
    bug_records.append(rec)

print("\nTotal Bug tickets:", len(bug_records))
print("Bug done:", sum(r["done"] for r in bug_records))
bug_priority_counts = {}
for r in bug_records:
    bug_priority_counts[r["priority"]] = bug_priority_counts.get(r["priority"], 0) + 1
print("Bug priority counts (all):", bug_priority_counts)
bug_not_done = [r for r in bug_records if not r["done"]]
bug_bucket_counts = {}
for r in bug_not_done:
    bug_bucket_counts[r["labelBucket"]] = bug_bucket_counts.get(r["labelBucket"], 0) + 1
print("Bug not-done label buckets:", bug_bucket_counts)
bug_feat_counts = {}
for r in bug_records:
    bug_feat_counts[r["feature"]] = bug_feat_counts.get(r["feature"], 0) + 1
print("Bug feature counts:", bug_feat_counts)

bug_team_counts = {}
for r in bug_not_done:
    bug_team_counts[r["team"]] = bug_team_counts.get(r["team"], 0) + 1
print("Bug not-done team counts:", dict(sorted(bug_team_counts.items(), key=lambda x: -x[1])))

bug_severity_counts = {}
for r in bug_records:
    bug_severity_counts[r["severity"]] = bug_severity_counts.get(r["severity"], 0) + 1
print("Bug severity counts (all):", dict(sorted(bug_severity_counts.items(), key=lambda x: -x[1])))
bug_unknown_assignees = sorted(set(r["assignee"] for r in bug_not_done if r["team"] == "Unknown"))
if bug_unknown_assignees:
    print(f"Bug assignees with NO team mapping ({len(bug_unknown_assignees)}), flagged as 'Unknown':")
    for a in bug_unknown_assignees:
        print("  ", a)


# --- Audio tickets (summary mentions "audio", any Issue Type) ---
def audio_group(summary, issue_type):
    if issue_type == "Bug":
        return "Bug"
    for tag in ("SWE2", "SWE3", "SWE5"):
        if tag in summary:
            return tag
    if "SWE1" in summary:
        return "SWE1"  # excluded below; not part of the tracked Audio breakdown
    return "Other"

# Assignees who work exclusively on audio -- include ALL of their tickets in the
# Audio tab even when the summary text itself doesn't contain "audio".
AUDIO_ASSIGNEES = {"KVVD Subrahmanyam", "Subramanya N S"}

# Tickets manually confirmed by the user as audio-related even though the summary
# text doesn't contain "audio" and the assignee isn't an audio-only assignee.
AUDIO_EXTRA_KEYS = {"NR1LT-3385"}

audio_records = []
for r in data:
    summary = r[IDX_SUMMARY] or ""
    assignee_name = (r[IDX_ASSIGNEE] if IDX_ASSIGNEE is not None else None) or "Unassigned"
    is_audio_summary = bool(re.search(r'audio', summary, re.I))
    is_audio_assignee = assignee_name in AUDIO_ASSIGNEES
    is_audio_extra = r[IDX_KEY] in AUDIO_EXTRA_KEYS
    if not is_audio_summary and not is_audio_assignee and not is_audio_extra:
        continue
    labels = labels_of(r)
    group = audio_group(summary, r[IDX_TYPE])
    if group == "SWE1":
        continue
    has_high_pri_dep = any("HighPriDep" in l for l in labels)
    # HighPriDep restriction applies only to the SWE2 group; SWE3/SWE5/Bug/others are unrestricted.
    if group == "SWE2" and not has_high_pri_dep:
        continue
    status = r[IDX_STATUS]
    cat = status_category(status)
    rec = {
        "key": r[IDX_KEY],
        "summary": summary,
        "issueType": r[IDX_TYPE],
        "status": status,
        "statusCategory": cat,
        "labels": labels,
        "group": group,
        "assignee": (r[IDX_ASSIGNEE] if IDX_ASSIGNEE is not None else None) or "Unassigned",
        "done": cat == "done",
        "hasR2": "ASW-R2" in labels,
        "hasR3": "ASW-R3" in labels,
        "hasCPAA0830": "CPAA_0830" in labels,
    }
    rec["labelBucket"] = label_bucket(rec["hasR2"], rec["hasR3"], rec["hasCPAA0830"])
    audio_records.append(rec)

print("\nTotal Audio tickets:", len(audio_records))
print("Audio done:", sum(r["done"] for r in audio_records))
audio_group_counts = {}
for r in audio_records:
    audio_group_counts[r["group"]] = audio_group_counts.get(r["group"], 0) + 1
print("Audio group counts:", audio_group_counts)

# --- Pretest tickets (Bug tickets whose title mentions PCTS or Facet) ---
pretest_records = []
for r in bug_records:
    s_lower = r["summary"].lower()
    has_facet = "facet" in s_lower
    has_pcts = "pcts" in s_lower
    if not has_facet and not has_pcts:
        continue
    rec = dict(r)
    rec["pretestGroup"] = "CP (Facet)" if has_facet else "AA (PCTS)"
    pretest_records.append(rec)

print("\nTotal Pretest tickets:", len(pretest_records))
pretest_group_counts = {}
for r in pretest_records:
    pretest_group_counts[r["pretestGroup"]] = pretest_group_counts.get(r["pretestGroup"], 0) + 1
print("Pretest group counts:", pretest_group_counts)

with open("dashboard_data.json", "w", encoding="utf-8") as f:
    json.dump({"tickets": records, "bugs": bug_records, "audio": audio_records, "pretest": pretest_records}, f, ensure_ascii=False, indent=1)
print("\nWrote dashboard_data.json")
